#!/usr/bin/env python3
"""Regression self-test for the two generator scripts.

Every case here corresponds to a defect that once shipped. Run it after any change
to either script, and before trusting a workbook you are about to send:

    python3 scripts/selftest.py

Formula results are checked by recalculating with LibreOffice when `soffice` is on
PATH. Without it, the arithmetic cases are reported as SKIPPED rather than silently
passing - a workbook whose totals were never recalculated has not really been tested.
"""
import os
import re
import shutil
import subprocess
import sys
import tempfile

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "build_baseline_xlsx.py")
DASH = os.path.join(HERE, "build_dashboard.py")
HAS_SOFFICE = shutil.which("soffice") is not None

results = []


def record(name, status, detail=""):
    results.append((name, status, detail))
    mark = {"PASS": "PASS", "FAIL": "FAIL", "SKIP": "SKIP"}[status]
    print(f"  [{mark}] {name}" + (f"  — {detail}" if detail else ""))


def run(args):
    return subprocess.run([sys.executable] + args, capture_output=True, text=True)


def totals(path, tmp):
    """Fill every seeded row with 2 x 30min = 1.0h, recalculate, return (total, unmatched)."""
    from openpyxl import load_workbook

    wb = load_workbook(path)
    ws = wb["Task Log"]
    filled = 0
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=2).value:
            ws.cell(row=r, column=4, value=2)
            ws.cell(row=r, column=5, value=30)
            filled += 1
    wb.save(path)
    # Convert into a separate directory: LibreOffice will not overwrite a file with
    # itself, so converting into the source directory silently leaves the formulas
    # uncalculated and every assertion below reads None.
    outdir = os.path.join(tmp, "recalc")
    os.makedirs(outdir, exist_ok=True)
    subprocess.run(["soffice", "--headless", "--convert-to", "xlsx", "--outdir", outdir, path],
                   capture_output=True, timeout=180)
    rec = os.path.join(outdir, os.path.basename(path))
    if not os.path.exists(rec):
        return filled, None, None, ["recalculation produced no file"]
    wb2 = load_workbook(rec, data_only=True)
    ro = wb2["Rollup"]
    total = unmatched = None
    errors = []
    for ws2 in wb2.worksheets:
        for row in ws2.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("#"):
                    errors.append(f"{ws2.title}!{c.coordinate}={c.value}")
    for i in range(2, 40):
        label = str(ro.cell(row=i, column=1).value or "")
        if label.startswith("TEAM TOTAL"):
            total = ro.cell(row=i, column=2).value
        elif label.startswith("UNMATCHED"):
            unmatched = ro.cell(row=i, column=2).value
    return filled, total, unmatched, errors




def check_frontmatter():
    """Refuse to pass if SKILL.md would be rejected on save.

    The platform caps `description` at 1024. It is not documented whether that
    counts characters or bytes, and these descriptions carry diacritics, so both
    are held under 1000. A skill that will not save is a skill that does not
    exist, and nothing else in this suite would have noticed.
    """
    import re as _re
    import pathlib as _pl
    md = (_pl.Path(__file__).resolve().parent.parent / "SKILL.md").read_text(encoding="utf-8")
    parts = md.split("---")
    problems = []
    if len(parts) < 3:
        problems.append("SKILL.md has no --- frontmatter block")
    else:
        fm = parts[1]
        if not _re.search(r"^name:\s*\S", fm, _re.M):
            problems.append("no name field")
        m = _re.search(r"^description:\s*(.*?)(?=^\w+:|\Z)", fm, _re.S | _re.M)
        if not m:
            problems.append("no description field")
        else:
            d = m.group(1).strip()
            if d.startswith(">"):
                d = " ".join(l.strip() for l in d.splitlines()[1:] if l.strip())
            if len(d) > 1000:
                problems.append(f"description is {len(d)} characters; the platform refuses over 1024")
            if len(d.encode("utf-8")) > 1000:
                problems.append(f"description is {len(d.encode('utf-8'))} bytes; diacritics cost extra")
    print("== frontmatter ==")
    for p in problems:
        print(f"  FAIL {p}")
    if problems:
        raise SystemExit(1)

def main():
    check_frontmatter()
    tmp = tempfile.mkdtemp(prefix="tep-selftest-")
    print(f"Workspace: {tmp}\n")

    print("Workbook — argument handling")
    for bad in ("0", "-3"):
        r = run([XLSX, "--client", "A", "--tasks", "a", "--output", f"{tmp}/x.xlsx",
                 "--rows-per-participant", bad])
        record(f"rejects --rows-per-participant {bad}",
               "PASS" if r.returncode != 0 and "at least 1" in r.stderr else "FAIL",
               r.stderr.strip().splitlines()[-1] if r.stderr else "")

    r = run([XLSX, "--client", "A", "--tasks", *[f"t{i}" for i in range(1, 8)],
             "--participants", "P1", "P2", "--output", f"{tmp}/many.xlsx"])
    record("raises row count so every task has a row plus spares",
           "PASS" if "rows/participant: 9 (7 seeded + 2 blank)" in r.stdout else "FAIL",
           [l for l in r.stdout.splitlines() if "rows/participant" in l][0].strip())

    r = run([XLSX, "--client", "A", "--tasks", "Izvestaj", "izvestaj", "Izvestaj ",
             "--participants", "P1", "--output", f"{tmp}/dupes.xlsx"])
    record("merges case/whitespace duplicate task names",
           "PASS" if r.stdout.count("MERGED:") == 2 else "FAIL",
           f"{r.stdout.count('MERGED:')} merges reported")

    print("\nWorkbook — arithmetic (requires LibreOffice recalculation)")
    if not HAS_SOFFICE:
        for n in ("wildcards do not absorb other rows", "no #DIV/0! before annoyance is scored",
                  "unmatched hours are surfaced, not lost"):
            record(n, "SKIP", "soffice not on PATH")
    else:
        run([XLSX, "--client", "A", "--tasks", "unos *", "unos podataka", "unos faktura",
             "--participants", "P1", "--output", f"{tmp}/wild.xlsx"])
        filled, total, unmatched, errors = totals(f"{tmp}/wild.xlsx", tmp)
        record("wildcards do not absorb other rows",
               "PASS" if total == filled else "FAIL", f"total={total}, expected {filled}")
        record("no #DIV/0! before annoyance is scored",
               "PASS" if not errors else "FAIL", "; ".join(errors[:3]) or "no error cells")

        run([XLSX, "--client", "A", "--tasks", "izvestaj", "ponude",
             "--participants", "P1", "--output", f"{tmp}/orphan.xlsx"])
        from openpyxl import load_workbook
        wb = load_workbook(f"{tmp}/orphan.xlsx")
        wb["Task Log"].cell(row=3, column=2, value="ponudee")  # deliberate typo
        wb.save(f"{tmp}/orphan.xlsx")
        filled, total, unmatched, _ = totals(f"{tmp}/orphan.xlsx", tmp)
        record("unmatched hours are surfaced, not lost",
               "PASS" if total == filled and unmatched and unmatched > 0 else "FAIL",
               f"total={total} (expected {filled}), unmatched={unmatched} (expected >0)")

    print("\nWorkbook — the before/after comparison must survive re-measurement")
    if not HAS_SOFFICE:
        for n in ("Week 0 and Week 12 measurements are kept in separate sheets",
                  "re-measuring does not erase the baseline",
                  "'people doing it' counts people who logged hours, not seeded rows",
                  "blank rows exist for tasks discovered during the week"):
            record(n, "SKIP", "soffice not on PATH")
    else:
        from openpyxl import load_workbook
        run([XLSX, "--client", "A", "--tasks", "alpha", "beta", "--participants", "P1", "P2",
             "--output", f"{tmp}/cmp.xlsx"])
        wb = load_workbook(f"{tmp}/cmp.xlsx")
        record("Week 0 and Week 12 measurements are kept in separate sheets",
               "PASS" if {"Task Log", "Week 12 Log"} <= set(wb.sheetnames) else "FAIL",
               str(wb.sheetnames))

        w0, w12 = wb["Task Log"], wb["Week 12 Log"]
        # Week 0: both people do alpha, 1h each. Week 12: same work, half the time.
        rows0 = [r for r in range(2, w0.max_row + 1) if w0.cell(row=r, column=2).value == "alpha"]
        for r in rows0:
            w0.cell(row=r, column=4, value=2)
            w0.cell(row=r, column=5, value=30)
        rows12 = [r for r in range(2, w12.max_row + 1) if w12.cell(row=r, column=2).value == "alpha"]
        for r in rows12:
            w12.cell(row=r, column=4, value=1)
            w12.cell(row=r, column=5, value=30)
        # One person logs nothing against beta, so "people doing it" must not count them.
        wb.save(f"{tmp}/cmp.xlsx")
        outdir = os.path.join(tmp, "recalc")
        os.makedirs(outdir, exist_ok=True)
        subprocess.run(["soffice", "--headless", "--convert-to", "xlsx", "--outdir", outdir,
                        f"{tmp}/cmp.xlsx"], capture_output=True, timeout=180)
        rec = load_workbook(os.path.join(outdir, "cmp.xlsx"), data_only=True)
        ro, cmp_ws = rec["Rollup"], rec["Week 12 Compare"]
        w0_alpha = ro.cell(row=2, column=2).value
        people_alpha = ro.cell(row=2, column=3).value
        people_beta = ro.cell(row=3, column=3).value
        c_before = c_after = c_delta = None
        for i in range(4, 12):
            if cmp_ws.cell(row=i, column=1).value == "alpha":
                c_before = cmp_ws.cell(row=i, column=2).value
                c_after = cmp_ws.cell(row=i, column=3).value
                c_delta = cmp_ws.cell(row=i, column=4).value
        record("re-measuring does not erase the baseline",
               "PASS" if (c_before == 2 and c_after == 1 and c_delta == 1) else "FAIL",
               f"before={c_before}, after={c_after}, delta={c_delta} (expected 2, 1, 1)")
        record("'people doing it' counts people who logged hours, not seeded rows",
               "PASS" if (people_alpha == 2 and people_beta == 0) else "FAIL",
               f"alpha={people_alpha} (expected 2), beta={people_beta} (expected 0)")
        blanks = sum(1 for r in range(2, w0.max_row + 1)
                     if w0.cell(row=r, column=1).value and not w0.cell(row=r, column=2).value)
        record("blank rows exist for tasks discovered during the week",
               "PASS" if blanks >= 2 else "FAIL", f"{blanks} blank rows")

    print("\nDashboard — parsing and failure modes")
    cases = {
        "bullets.md": "## Week 3 — Context\n- **Objective:** Stop re-explaining.\n"
                      "- **Exercise:** Set up instructions.\n- **Deliverable:** A workspace.\n",
        "serbian.md": "## Nedelja 5 — Prva vestina\n**Cilj.** Pretvoriti zadatak.\n"
                      "**Vezba.** Zapisati odluke.\n**Isporuka.** Jedna vestina.\n",
        "nofields.md": "## Week 4 — Tools\nProse with no labelled fields at all.\n",
        "dupes.md": "## Week 5 — First\n**Objective.** Keep me.\n"
                    "## Week 5 — Second\n**Objective.** Discard me.\n",
    }
    for name, body in cases.items():
        open(os.path.join(tmp, name), "w", encoding="utf-8").write(body)

    r = run([DASH, "--program", f"{tmp}/bullets.md", "--client", "T", "--output", f"{tmp}/b.html"])
    html = open(f"{tmp}/b.html", encoding="utf-8").read() if os.path.exists(f"{tmp}/b.html") else ""
    record("parses list-style field labels",
           "PASS" if "Stop re-explaining." in html else "FAIL")

    r = run([DASH, "--program", f"{tmp}/serbian.md", "--client", "T", "--output", f"{tmp}/s.html"])
    html = open(f"{tmp}/s.html", encoding="utf-8").read() if os.path.exists(f"{tmp}/s.html") else ""
    record("parses Serbian field labels",
           "PASS" if "Pretvoriti zadatak." in html else "FAIL")

    r = run([DASH, "--program", f"{tmp}/nofields.md", "--client", "T", "--output", f"{tmp}/n.html"])
    record("warns when a week has prose but no runnable field labels",
           "PASS" if "no Objective / Exercise / Deliverable" in r.stdout else "FAIL",
           r.stdout.strip().splitlines()[-1][:80] if r.stdout else "")

    open(f"{tmp}/nothing.md", "w", encoding="utf-8").write("## Week 4 — Tools\n\n| a | b |\n")
    r = run([DASH, "--program", f"{tmp}/nothing.md", "--client", "T", "--output", f"{tmp}/z.html"])
    record("refuses to emit an all-blank dashboard",
           "PASS" if r.returncode != 0 and not os.path.exists(f"{tmp}/z.html") else "FAIL",
           f"exit={r.returncode}")

    r = run([DASH, "--program", f"{tmp}/dupes.md", "--client", "T", "--output", f"{tmp}/d.html"])
    html = open(f"{tmp}/d.html", encoding="utf-8").read() if os.path.exists(f"{tmp}/d.html") else ""
    record("duplicate week keeps the first and warns",
           "PASS" if ("Keep me." in html and "Discard me." not in html
                      and "appear more than once" in r.stdout) else "FAIL")

    prose = ("## Week 4 — Tools\n**Objective.** Connect the stack.\n\n"
             "WhatsApp is deliberately not connected in any week of this programme.\n\n"
             "## Conditions of engagement\nThe sponsor attends every week.\n")
    open(f"{tmp}/prose.md", "w", encoding="utf-8").write(prose)
    run([DASH, "--program", f"{tmp}/prose.md", "--client", "T", "--output", f"{tmp}/p.html"])
    html = open(f"{tmp}/p.html", encoding="utf-8").read()
    record("restrictions written as prose survive into the dashboard",
           "PASS" if "WhatsApp is deliberately not connected" in html else "FAIL")
    record("rules and scope sections survive into the dashboard",
           "PASS" if "sponsor attends every week" in html.lower() else "FAIL")

    inj = "## Week 1 — Probe\n**Objective.** </script><script>alert(1)</script> & <b>x</b>\n"
    open(f"{tmp}/inj.md", "w", encoding="utf-8").write(inj)
    run([DASH, "--program", f"{tmp}/inj.md", "--client", "T", "--output", f"{tmp}/i.html"])
    html = open(f"{tmp}/i.html", encoding="utf-8").read()
    record("escapes HTML in field values",
           "PASS" if "<script>alert(1)</script>" not in html and "&lt;script&gt;" in html else "FAIL")

    failed = [n for n, s, _ in results if s == "FAIL"]
    skipped = [n for n, s, _ in results if s == "SKIP"]
    print(f"\n{'=' * 66}")
    print(f"{len(results) - len(failed) - len(skipped)} passed, {len(failed)} failed, "
          f"{len(skipped)} skipped")
    if failed:
        for n in failed:
            print(f"  FAILED: {n}")
    print("=" * 66)
    shutil.rmtree(tmp, ignore_errors=True)
    return 1 if failed else 0


if __name__ == "__main__":
    sys.exit(main())
