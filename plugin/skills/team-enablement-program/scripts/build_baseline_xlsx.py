#!/usr/bin/env python3
"""Build the Week 0 baseline measurement workbook for a team enablement program.

Five sheets:
  1. Read Me         - what this is for and how to fill it in
  2. Task Log        - Week 0 measurement, one row per person per task
  3. Week 12 Log     - the SAME form, filled at the end. Separate on purpose.
  4. Rollup          - Week 0 hours by task, plus an unmatched-hours guard
  5. Week 12 Compare - before/after, computed from the two logs

## Why Week 12 gets its own sheet

The obvious design - one log, re-measured at the end - destroys the baseline the
moment it is re-measured, because every total is a live formula over that one sheet.
The before column then equals the after column and the delta reads zero. The whole
commercial case rests on that comparison, so the two measurements live in two sheets
and neither can overwrite the other.

## Why totals use EXACT() rather than SUMIF()

SUMIF is case-insensitive and treats * and ? as wildcards. A task called
"unos *svih* podataka" would silently absorb every other row starting with "unos",
and "Izvestaj" would merge with "izvestaj" - both inflating the number the whole
programme is judged on. SUMPRODUCT(--EXACT(...)) matches literally.

Team totals are summed straight from the log rows rather than from the task rows, so
they stay correct even when a task name is mistyped, and any hours that match no
listed task surface in an UNMATCHED row instead of disappearing.

Usage:
  python3 build_baseline_xlsx.py --client "Acme" --team "Sales" \
      --tasks "qualify inbound leads" "write follow-up emails" \
      --participants "Ana Maric" "Marko Petrovic" \
      --output acme-week0-baseline.xlsx
"""

import argparse
import sys

try:
    from openpyxl import Workbook
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("openpyxl is required:  pip install openpyxl --break-system-packages")


HEADER_FILL = PatternFill("solid", fgColor="1F3A5F")
HEADER_FILL_LATE = PatternFill("solid", fgColor="3E5C3A")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F3A5F")
NOTE_FONT = Font(italic=True, size=9, color="666666")
CALC_FILL = PatternFill("solid", fgColor="EEF2F7")
INPUT_FILL = PatternFill("solid", fgColor="FFFDF5")
ALERT_FILL = PatternFill("solid", fgColor="FFD9D9")
ALERT_FONT = Font(color="9C1B1B", bold=True)
THIN = Side(style="thin", color="C9D2DC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

LOG_COLUMNS = [
    ("Participant", 22),
    ("Task (copy the name from Rollup)", 38),
    ("Tool used", 20),
    ("Times per week", 14),
    ("Minutes per occurrence", 20),
    ("Hours per week", 15),
    ("Who else does this", 24),
    ("Annoyance 1-5", 14),
    ("Notes", 34),
]
# Hidden helpers. They exist so the Rollup can use plain numeric ranges: the visible
# Hours column returns "" when empty, and text inside a SUMPRODUCT range raises
# #VALUE!, while a blank Annoyance column made AVERAGEIF raise #DIV/0!.
# _has_hours exists so "people doing it" counts people who actually logged time
# rather than pre-seeded rows, which always returned the full team size.
HELPERS = [("_hrs_num", 10), ("_annoy_val", 11), ("_annoy_cnt", 12), ("_has_hours", 13)]

COMPARE_COLUMNS = [
    ("Task", 38),
    ("Week 0 hrs/week", 16),
    ("Week 12 hrs/week", 17),
    ("Delta (hrs)", 13),
    ("Change %", 12),
    ("Verdict (keep/fix/kill)", 22),
    ("Evidence / notes", 40),
]

SPARE_ROWS = 2  # blank rows per participant for tasks discovered during the week


def style_header(ws, columns, row=1, fill=HEADER_FILL):
    for idx, (name, width) in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=idx, value=name)
        cell.fill = fill
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.row_dimensions[row].height = 30
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def dedupe_tasks(tasks):
    """Collapse names differing only by case or surrounding whitespace.

    Two Rollup rows with the same name each total the same log rows, and the team
    total added both - reporting 5 hours where 3 were worked.
    """
    seen, out, dropped = {}, [], []
    for t in tasks:
        stripped = t.strip()
        key = stripped.casefold()
        if key in seen:
            dropped.append((t, seen[key]))
            continue
        seen[key] = stripped
        out.append(stripped)
    return out, dropped


def build_log(wb, title, participants, tasks, rows_per_participant, late=False):
    ws = wb.create_sheet(title)
    style_header(ws, LOG_COLUMNS, fill=HEADER_FILL_LATE if late else HEADER_FILL)
    for name, idx in HELPERS:
        ws.cell(row=1, column=idx, value=name).font = Font(size=8, color="AAAAAA")
        ws.column_dimensions[get_column_letter(idx)].hidden = True

    row = 2
    for person in participants:
        for i in range(rows_per_participant):
            ws.cell(row=row, column=1, value=person)
            if i < len(tasks):
                ws.cell(row=row, column=2, value=tasks[i])
            ws.cell(row=row, column=6,
                    value=f"=IF(OR(D{row}=\"\",E{row}=\"\"),\"\",ROUND(D{row}*E{row}/60,2))")
            ws.cell(row=row, column=6).fill = CALC_FILL
            ws.cell(row=row, column=10, value=f"=IF(OR(D{row}=\"\",E{row}=\"\"),0,D{row}*E{row}/60)")
            ws.cell(row=row, column=11, value=f"=IF(ISNUMBER(H{row}),H{row},0)")
            ws.cell(row=row, column=12, value=f"=IF(ISNUMBER(H{row}),1,0)")
            ws.cell(row=row, column=13, value=f"=IF(J{row}>0,1,0)")
            for col in (2, 4, 5, 8):
                ws.cell(row=row, column=col).fill = INPUT_FILL
            for col in range(1, len(LOG_COLUMNS) + 1):
                ws.cell(row=row, column=col).border = BORDER
            row += 1

    last_row = row - 1

    dv = DataValidation(type="whole", operator="between", formula1=1, formula2=5,
                        allow_blank=True, showErrorMessage=True)
    dv.error = "Annoyance is a 1-5 rating."
    ws.add_data_validation(dv)
    dv.add(f"H2:H{last_row}")

    n = last_row + 2
    if late:
        ws.cell(row=n, column=1,
                value="Fill this in at Week 12, using exactly the method used in Week 0. Mixing an "
                      "estimated 'before' with a tracked 'after' manufactures an improvement that is "
                      "not there.").font = NOTE_FONT
        ws.cell(row=n + 1, column=1,
                value="The Week 0 sheet is left untouched on purpose - overwrite it and the "
                      "comparison has nothing to compare against.").font = NOTE_FONT
    else:
        ws.cell(row=n, column=1,
                value="Estimates are fine. Whatever method is used here must be used identically at "
                      "Week 12.").font = NOTE_FONT
        ws.cell(row=n + 1, column=1,
                value=f"Blank rows are for tasks discovered during the week. Use them - rows added "
                      f"below row {last_row} fall outside every total on the Rollup. Add any new task "
                      f"to the Rollup list as well, or its hours land in UNMATCHED.").font = NOTE_FONT
        ws.cell(row=n + 2, column=1,
                value="Task names are matched literally, including capitalisation. Copy the name from "
                      "the Rollup sheet rather than retyping it.").font = NOTE_FONT
    return last_row


def build_rollup(wb, tasks, log_last, spare_note_rows):
    ws = wb.create_sheet("Rollup")
    columns = [("Task", 40), ("Total hrs/week (team)", 22), ("People logging hours", 20),
               ("Avg annoyance", 15), ("Priority note (fill in when reviewing)", 40)]
    style_header(ws, columns)

    task_r = f"'Task Log'!$B$2:$B${log_last}"
    hrs_r = f"'Task Log'!$J$2:$J${log_last}"
    aval_r = f"'Task Log'!$K$2:$K${log_last}"
    acnt_r = f"'Task Log'!$L$2:$L${log_last}"
    has_r = f"'Task Log'!$M$2:$M${log_last}"

    seed = tasks if tasks else [""] * 8
    seed = seed + [""] * spare_note_rows  # blank rows for tasks discovered in Week 0
    row = 2
    for task in seed:
        ws.cell(row=row, column=1, value=task)
        ws.cell(row=row, column=1).fill = INPUT_FILL
        m = f"--EXACT({task_r},$A{row})"
        ws.cell(row=row, column=2,
                value=f"=IF($A{row}=\"\",\"\",ROUND(SUMPRODUCT({m},{hrs_r}),2))")
        # Count people who logged time, not pre-seeded rows. The old version returned
        # the full team size for every task whether anyone did it or not.
        ws.cell(row=row, column=3,
                value=f"=IF($A{row}=\"\",\"\",SUMPRODUCT({m},{has_r}))")
        ws.cell(row=row, column=4,
                value=f"=IF($A{row}=\"\",\"\",IF(SUMPRODUCT({m},{acnt_r})=0,\"\","
                      f"ROUND(SUMPRODUCT({m},{aval_r})/SUMPRODUCT({m},{acnt_r}),1)))")
        for col in (2, 3, 4):
            ws.cell(row=row, column=col).fill = CALC_FILL
        ws.cell(row=row, column=5).fill = INPUT_FILL
        for col in range(1, len(columns) + 1):
            ws.cell(row=row, column=col).border = BORDER
        row += 1

    last_task = row - 1
    total_row, unmatched_row = row + 1, row + 2

    ws.cell(row=total_row, column=1, value="TEAM TOTAL (from Task Log)").font = Font(bold=True)
    ws.cell(row=total_row, column=2, value=f"=ROUND(SUM({hrs_r}),2)").font = Font(bold=True)
    ws.cell(row=unmatched_row, column=1,
            value="UNMATCHED hours (rows matching no task above)").font = Font(bold=True)
    ws.cell(row=unmatched_row, column=2,
            value=f"=ROUND(B{total_row}-SUM(B2:B{last_task}),2)").font = Font(bold=True)
    ws.conditional_formatting.add(
        f"B{unmatched_row}",
        CellIsRule(operator="greaterThan", formula=["0.001"], fill=ALERT_FILL, font=ALERT_FONT))

    ws.cell(row=unmatched_row + 2, column=1,
            value=f"A non-zero UNMATCHED figure means somebody logged a task name that is not listed "
                  f"above - usually a typo, sometimes a task discovered during the week. Type the new "
                  f"task into one of the blank rows between 2 and {last_task}. Do not insert a row "
                  f"below {last_task}: the totals do not reach it.").font = NOTE_FONT
    ws.auto_filter.ref = f"A1:E{last_task}"
    ws.cell(row=unmatched_row + 4, column=1,
            value="Use the filter arrows on row 1 to sort by total hours, descending. Week 5's first "
                  "skill targets a task near the top - a first skill saving twenty minutes a week "
                  "does not survive a busy month.").font = NOTE_FONT
    ws.cell(row=unmatched_row + 5, column=1,
            value="Cross-check against annoyance: high hours with low annoyance usually meets "
                  "resistance, mid hours with high annoyance usually meets enthusiasm.").font = NOTE_FONT

    s = unmatched_row + 7
    ws.cell(row=s, column=1, value="Reviewed and agreed by (name):").font = Font(bold=True)
    ws.cell(row=s + 1, column=1, value="Date:").font = Font(bold=True)
    for r in (s, s + 1):
        ws.cell(row=r, column=2).fill = INPUT_FILL
        ws.cell(row=r, column=2).border = BORDER
    ws.cell(row=s + 3, column=1,
            value="Agreeing the baseline now is what makes the Week 12 comparison meaningful to "
                  "everyone in the room.").font = NOTE_FONT


def build_compare(wb, tasks, late_last, spare_note_rows):
    ws = wb.create_sheet("Week 12 Compare")
    ws.cell(row=1, column=1,
            value="Fills itself once the Week 12 Log is complete. Nothing to type except the "
                  "verdict.").font = TITLE_FONT
    style_header(ws, COMPARE_COLUMNS, row=3)

    late_task = f"'Week 12 Log'!$B$2:$B${late_last}"
    late_hrs = f"'Week 12 Log'!$J$2:$J${late_last}"

    seed = (tasks if tasks else [""] * 8) + [""] * spare_note_rows
    row = 4
    for task in seed:
        ws.cell(row=row, column=1,
                value=f"=IF(Rollup!A{row - 2}=\"\",\"\",Rollup!A{row - 2})")
        ws.cell(row=row, column=2,
                value=f"=IF($A{row}=\"\",\"\",IFERROR(VLOOKUP($A{row},Rollup!$A:$B,2,FALSE),\"\"))")
        ws.cell(row=row, column=3,
                value=f"=IF($A{row}=\"\",\"\",ROUND(SUMPRODUCT(--EXACT({late_task},$A{row}),"
                      f"{late_hrs}),2))")
        ws.cell(row=row, column=4,
                value=f"=IF(OR(B{row}=\"\",C{row}=\"\"),\"\",ROUND(B{row}-C{row},2))")
        ws.cell(row=row, column=5,
                value=f"=IF(OR(B{row}=\"\",C{row}=\"\",B{row}=0),\"\",ROUND((B{row}-C{row})/B{row}*100,1))")
        for col in (1, 2, 3, 4, 5):
            ws.cell(row=row, column=col).fill = CALC_FILL
        for col in (6, 7):
            ws.cell(row=row, column=col).fill = INPUT_FILL
        for col in range(1, len(COMPARE_COLUMNS) + 1):
            ws.cell(row=row, column=col).border = BORDER
        row += 1

    last = row - 1
    ws.conditional_formatting.add(
        f"D4:E{last}",
        CellIsRule(operator="lessThan", formula=["0"], fill=ALERT_FILL, font=ALERT_FONT))

    ws.cell(row=row + 1, column=1, value="Report three numbers, not one:").font = Font(bold=True)
    for i, t in enumerate(["1. Hours saved per week (sum of the Delta column)",
                           "2. Tasks that improved, out of tasks attempted",
                           "3. Artefacts killed at the Week 12 review"]):
        ws.cell(row=row + 2 + i, column=1, value=t)
    ws.cell(row=row + 6, column=1,
            value="A positive Delta and a positive Change % mean hours went DOWN. Rows shaded red "
                  "went the wrong way - hours went up. That is a real finding, not an embarrassment: "
                  "a skill that added work is a skill to kill.").font = NOTE_FONT
    ws.cell(row=row + 7, column=1,
            value="Both columns are self-reported by the same people, so this measures perceived "
                  "time, not stopwatch time. Say so when presenting it. The comparison is honest "
                  "because the method is identical at both ends, not because it is precise.").font = NOTE_FONT


def build_cover(wb, client, team, participants, tasks):
    ws = wb.active
    ws.title = "Read Me"
    ws.column_dimensions["A"].width = 100
    lines = [
        (f"{client} - Week 0 Baseline", TITLE_FONT),
        (f"Team: {team}    Participants: {len(participants)}", Font(size=11)),
        ("", None),
        ("What this is for", Font(bold=True, size=12)),
        ("Choosing which tasks to attack first, and being able to show at the end whether it", None),
        ("worked. Without a before number, any claim about time saved is an opinion.", None),
        ("", None),
        ("How to fill it in", Font(bold=True, size=12)),
        ("1. Each participant logs their most repeated tasks in the Task Log sheet.", None),
        ("2. Log at the end of each day for one representative working week. Reconstructing the", None),
        ("   week on Friday tends to underestimate small repeated tasks - the exact ones this", None),
        ("   programme targets.", None),
        ("3. Estimates are fine. Consistency of method matters more than precision.", None),
        ("4. Copy task names from the Rollup sheet rather than retyping them. Names are matched", None),
        ("   literally, so a typo moves those hours into the UNMATCHED row.", None),
        ("5. Found a task that is not listed? Use one of the blank rows, and add the same name to", None),
        ("   a blank row on the Rollup. Rows added below the block are outside the totals.", None),
        ("6. Review the Rollup together and agree it before Week 1 begins.", None),
        ("", None),
        ("At Week 12", Font(bold=True, size=12)),
        ("Fill in the Week 12 Log sheet - the same form again, same method. The Week 0 sheet stays", None),
        ("untouched, and the comparison fills itself in.", None),
        ("", None),
        ("What not to do", Font(bold=True, size=12)),
        ("This is not a productivity audit and should not be presented as one. If people think the", None),
        ("numbers will be used to judge them individually, the numbers stop being useful. Before", None),
        ("collecting anything, participants should be told in writing what is collected, who sees", None),
        ("it, how long it is kept, and that it will not be used in performance assessment.", None),
        ("", None),
        ("Sheets", Font(bold=True, size=12)),
        ("Read Me          - this page", None),
        ("Task Log         - fill in during Week 0", None),
        ("Week 12 Log      - fill in at Week 12, same form", None),
        ("Rollup           - Week 0 totals by task", None),
        ("Week 12 Compare  - before and after, fills itself", None),
    ]
    for i, (text, font) in enumerate(lines, start=1):
        c = ws.cell(row=i, column=1, value=text)
        if font:
            c.font = font
    if tasks:
        s = len(lines) + 2
        ws.cell(row=s, column=1, value="Repeated tasks supplied at intake").font = Font(bold=True, size=12)
        for j, t in enumerate(tasks, start=1):
            ws.cell(row=s + j, column=1, value=f"  - {t}")


def main():
    p = argparse.ArgumentParser(description="Build the Week 0 baseline workbook.")
    p.add_argument("--client", required=True)
    p.add_argument("--team", default="Team")
    p.add_argument("--tasks", nargs="*", default=[])
    p.add_argument("--participants", nargs="*", default=[])
    p.add_argument("--rows-per-participant", type=int, default=5,
                   help="Task rows per participant. Raised automatically so every intake task has "
                        "a row plus blank rows for discovery (default 5).")
    p.add_argument("--output", required=True)
    args = p.parse_args()

    if args.rows_per_participant < 1:
        p.error("--rows-per-participant must be at least 1")

    tasks, dropped = dedupe_tasks(args.tasks)
    participants = args.participants or [f"Participant {i}" for i in range(1, 6)]

    # Every intake task needs a row, and the week needs room for what it turns up.
    rows = max(args.rows_per_participant, len(tasks) + SPARE_ROWS)

    wb = Workbook()
    build_cover(wb, args.client, args.team, participants, tasks)
    last = build_log(wb, "Task Log", participants, tasks, rows)
    late_last = build_log(wb, "Week 12 Log", participants, tasks, rows, late=True)
    build_rollup(wb, tasks, last, SPARE_ROWS)
    build_compare(wb, tasks, late_last, SPARE_ROWS)
    wb.save(args.output)

    print(f"Wrote {args.output}")
    print(f"  participants: {len(participants)}   rows/participant: {rows} "
          f"({len(tasks)} seeded + {rows - len(tasks)} blank)   log rows: {last - 1} per sheet")
    for orig, kept in dropped:
        print(f"  MERGED: {orig!r} duplicates {kept!r} (case/whitespace) - kept one row")
    if rows > args.rows_per_participant:
        print(f"  NOTE: raised rows per participant from {args.rows_per_participant} to {rows} so "
              f"every intake task has a row and {SPARE_ROWS} stay blank for discovery.")
    if not tasks:
        print("  NOTE: no tasks supplied - Rollup and Compare are blank templates.")


if __name__ == "__main__":
    main()
